from pathlib import Path
from datetime import date
from numbers import Number
import math
import statistics

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# SETTINGS
# ============================================================

INPUT_XLSX = Path("Score_Class_2026.xlsx")
OUTPUT_XLSX = Path("Score_Class_2026_Summary.xlsx")
OUTPUT_MD = Path("Score_Summary.md")
OUTPUT_HISTOGRAM = Path("Grand_Total_Histogram.png")

EXERCISE_SHEET = "Exercise"
ATTENDANCE_SHEET = "ClassAttend"
SUMMARY_SHEET = "Summary"

FIRST_STUDENT_ROW = 8

# Exercise sheet
EXERCISE_ID_COLUMN = 3       # C
EXERCISE_NAME_COLUMN = 4     # D
EXERCISE_START_COLUMN = 12   # L
EXERCISE_TOTAL_ROW = 7
EXERCISE_WEIGHT = 70.0

# Attendance sheet
ATTENDANCE_ID_COLUMN = 2      # B
ATTENDANCE_START_COLUMN = 4   # D
ATTENDANCE_HEADER_ROW = 6
ATTENDANCE_WEIGHT = 30.0


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def is_number(value):
    """Return True for valid numeric values, excluding booleans."""
    return isinstance(value, Number) and not isinstance(value, bool)


def number_value(value):
    """Convert a cell value to float; blanks and non-numbers become zero."""
    if is_number(value):
        return float(value)

    if isinstance(value, str):
        text = value.strip()

        if not text:
            return 0.0

        try:
            return float(text)
        except ValueError:
            return 0.0

    return 0.0


def normalize_student_id(value):
    """Convert student ID to a clean string."""
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def is_attended(value):
    """
    Interpret common attendance values.

    Attended:
      1, positive numbers, yes, present, attended, x, check marks

    Not attended:
      blank, 0, absent, no
    """
    if value is None:
        return False

    if is_number(value):
        return float(value) > 0

    text = str(value).strip().lower()

    attended_values = {
        "1",
        "yes",
        "y",
        "present",
        "attended",
        "มา",
        "เข้าเรียน",
        "x",
        "✓",
        "✔",
    }

    return text in attended_values


def star_rating(grand_total):
    """
    Performance evaluation:
       90–100 = 5 stars
       80–89  = 4 stars
       70–79  = 3 stars
       60–69  = 2 stars
       below 60 = 1 star
    """
    if grand_total >= 90:
        count = 5
    elif grand_total >= 80:
        count = 4
    elif grand_total >= 70:
        count = 3
    elif grand_total >= 60:
        count = 2
    else:
        count = 1

    return "★" * count + "☆" * (5 - count)


def markdown_text(value):
    """
    Escape Markdown-sensitive characters and use <sub>
    to display compact text on GitHub.
    """
    text = "" if value is None else str(value)
    text = text.replace("|", r"\|")
    text = text.replace("\n", " ")
    return f"<sub>{text}</sub>"


def find_last_exercise_column(ws):
    """
    Find the last exercise column.

    An exercise column is detected when row 7 contains
    a positive numeric maximum score.
    """
    exercise_columns = []

    for col in range(EXERCISE_START_COLUMN, ws.max_column + 1):
        maximum_score = number_value(
            ws.cell(EXERCISE_TOTAL_ROW, col).value
        )

        if maximum_score > 0:
            exercise_columns.append(col)

    if not exercise_columns:
        raise ValueError(
            "No exercise maximum scores were found in row 7 "
            "from column L onward."
        )

    return exercise_columns


def find_attendance_columns(ws):
    """
    Find active attendance columns from column D onward.

    A column is active when:
    1. Row 6 contains a date/header, or
    2. At least one student row contains attendance information.
    """
    attendance_columns = []

    for col in range(ATTENDANCE_START_COLUMN, ws.max_column + 1):
        header = ws.cell(ATTENDANCE_HEADER_ROW, col).value

        has_header = header is not None and str(header).strip() != ""

        has_student_data = any(
            ws.cell(row, col).value not in (None, "")
            for row in range(FIRST_STUDENT_ROW, ws.max_row + 1)
        )

        if has_header or has_student_data:
            attendance_columns.append(col)

    if not attendance_columns:
        raise ValueError(
            "No attendance columns were found from column D onward."
        )

    return attendance_columns


def resolve_attendance_student_id(cell_value, exercise_ws):
    """
    Resolve student ID from ClassAttend.

    ClassAttend may contain:
      - a direct student ID, or
      - a formula such as =Exercise!C8
    """
    if cell_value is None:
        return ""

    if isinstance(cell_value, str) and cell_value.startswith("="):
        formula = cell_value.replace("$", "").strip()

        # Expected example: =Exercise!C8
        if "!" in formula:
            reference = formula.split("!", 1)[1]

            row_digits = "".join(
                character
                for character in reference
                if character.isdigit()
            )

            if row_digits:
                exercise_row = int(row_digits)
                return normalize_student_id(
                    exercise_ws.cell(
                        exercise_row,
                        EXERCISE_ID_COLUMN
                    ).value
                )

    return normalize_student_id(cell_value)


def create_grand_total_histogram(grand_totals, output_path):
    """
    Create a colourful histogram of Grand Total scores.

    Returns:
      mean_score, standard_deviation, total_students

    The standard deviation is the population SD because the results
    represent the complete class rather than a sample of the class.
    """
    if not grand_totals:
        raise ValueError("No Grand Total scores are available.")

    total_students = len(grand_totals)
    mean_score = statistics.mean(grand_totals)
    standard_deviation = statistics.pstdev(grand_totals)

    # Use 10-point bins. Keep the normal 0–100 score range visible,
    # but expand automatically if an unexpected score is outside it.
    lower_limit = min(0, math.floor(min(grand_totals) / 10) * 10)
    upper_limit = max(100, math.ceil(max(grand_totals) / 10) * 10)

    if upper_limit == lower_limit:
        upper_limit = lower_limit + 10

    bin_edges = list(range(lower_limit, upper_limit + 10, 10))

    figure, axis = plt.subplots(figsize=(10, 5.8), dpi=160)

    counts, _, patches = axis.hist(
        grand_totals,
        bins=bin_edges,
        edgecolor="white",
        linewidth=1.2,
        rwidth=0.92,
    )

    colour_map = plt.get_cmap("turbo")
    colour_count = max(1, len(patches) - 1)

    for index, patch in enumerate(patches):
        patch.set_facecolor(colour_map(index / colour_count))

    # Display the number of students above every non-empty bin.
    top_count = max(counts) if len(counts) else 0
    label_offset = max(0.15, top_count * 0.025)

    for count, patch in zip(counts, patches):
        if count > 0:
            axis.text(
                patch.get_x() + patch.get_width() / 2,
                count + label_offset,
                f"{int(count)}",
                ha="center",
                va="bottom",
                fontsize=9,
                fontweight="bold",
            )

    axis.axvline(
        mean_score,
        color="black",
        linestyle="--",
        linewidth=2.0,
        label=f"Mean = {mean_score:.2f}",
    )

    if standard_deviation > 0:
        axis.axvspan(
            mean_score - standard_deviation,
            mean_score + standard_deviation,
            color="black",
            alpha=0.08,
            label=f"Mean ± 1 SD ({standard_deviation:.2f})",
        )

    current_date = f"{date.today().day} {date.today().strftime('%b %Y')}"

    axis.set_title(
        f"Score Distribution until {current_date} !!!",
        fontsize=15,
        fontweight="bold",
        color="red",
        pad=30,
    )
    axis.text(
        0.5,
        1.015,
        (
            f"Students: {total_students}   |   "
            f"Mean: {mean_score:.2f}   |   "
            f"SD: {standard_deviation:.2f} (population)"
        ),
        transform=axis.transAxes,
        ha="center",
        va="bottom",
        fontsize=10,
    )
    axis.set_xlabel("Grand Total score", fontsize=11)
    axis.set_ylabel("Number of students", fontsize=11)
    axis.set_xticks(bin_edges)
    axis.set_xlim(bin_edges[0], bin_edges[-1])
    axis.set_ylim(bottom=0)
    axis.grid(axis="y", linestyle=":", alpha=0.45)
    axis.legend(loc="upper left", frameon=True)

    figure.tight_layout()
    figure.savefig(output_path, bbox_inches="tight")
    plt.close(figure)

    return mean_score, standard_deviation, total_students


# ============================================================
# MAIN PROCESS
# ============================================================

def main():
    if not INPUT_XLSX.exists():
        raise FileNotFoundError(
            f"Input workbook not found: {INPUT_XLSX.resolve()}"
        )

    # data_only=False is required because ClassAttend contains
    # formulas linking student IDs to the Exercise worksheet.
    workbook = load_workbook(INPUT_XLSX, data_only=False)

    if EXERCISE_SHEET not in workbook.sheetnames:
        raise KeyError(
            f"Worksheet '{EXERCISE_SHEET}' was not found."
        )

    if ATTENDANCE_SHEET not in workbook.sheetnames:
        raise KeyError(
            f"Worksheet '{ATTENDANCE_SHEET}' was not found."
        )

    exercise_ws = workbook[EXERCISE_SHEET]
    attendance_ws = workbook[ATTENDANCE_SHEET]

    exercise_columns = find_last_exercise_column(exercise_ws)
    attendance_columns = find_attendance_columns(attendance_ws)

    total_exercise_possible = sum(
        number_value(
            exercise_ws.cell(EXERCISE_TOTAL_ROW, col).value
        )
        for col in exercise_columns
    )

    total_attendance_days = len(attendance_columns)

    if total_exercise_possible <= 0:
        raise ValueError("Total possible exercise score is zero.")

    if total_attendance_days <= 0:
        raise ValueError("Total attendance days is zero.")

    # --------------------------------------------------------
    # Create attendance lookup by Student ID
    # --------------------------------------------------------

    attendance_lookup = {}

    for row in range(FIRST_STUDENT_ROW, attendance_ws.max_row + 1):
        student_id = resolve_attendance_student_id(
            attendance_ws.cell(
                row,
                ATTENDANCE_ID_COLUMN
            ).value,
            exercise_ws
        )

        if not student_id:
            continue

        attended_count = sum(
            1
            for col in attendance_columns
            if is_attended(attendance_ws.cell(row, col).value)
        )

        attendance_lookup[student_id] = attended_count

    # --------------------------------------------------------
    # Calculate student scores
    # --------------------------------------------------------

    results = []

    for row in range(FIRST_STUDENT_ROW, exercise_ws.max_row + 1):
        student_id = normalize_student_id(
            exercise_ws.cell(
                row,
                EXERCISE_ID_COLUMN
            ).value
        )

        student_name = exercise_ws.cell(
            row,
            EXERCISE_NAME_COLUMN
        ).value

        if not student_id:
            continue

        raw_exercise_score = sum(
            number_value(exercise_ws.cell(row, col).value)
            for col in exercise_columns
        )

        exercise_percent = (
            raw_exercise_score / total_exercise_possible
        ) * 100.0

        weighted_exercise_score = (
            raw_exercise_score / total_exercise_possible
        ) * EXERCISE_WEIGHT

        attended_count = attendance_lookup.get(student_id, 0)

        attendance_percent = (
            attended_count / total_attendance_days
        ) * 100.0

        weighted_attendance_score = (
            attended_count / total_attendance_days
        ) * ATTENDANCE_WEIGHT

        grand_total = (
            weighted_exercise_score +
            weighted_attendance_score
        )

        results.append({
            "student_id": student_id,
            "student_name": student_name or "",
            "exercise_raw": raw_exercise_score,
            "exercise_possible": total_exercise_possible,
            "exercise_percent": exercise_percent,
            "exercise_weighted": weighted_exercise_score,
            "attended": attended_count,
            "attendance_days": total_attendance_days,
            "attendance_percent": attendance_percent,
            "attendance_weighted": weighted_attendance_score,
            "grand_total": grand_total,
            "stars": star_rating(grand_total),
        })

    grand_totals = [result["grand_total"] for result in results]

    mean_score, standard_deviation, total_students = (
        create_grand_total_histogram(
            grand_totals,
            OUTPUT_HISTOGRAM,
        )
    )

    # --------------------------------------------------------
    # Create Summary worksheet
    # --------------------------------------------------------

    if SUMMARY_SHEET in workbook.sheetnames:
        del workbook[SUMMARY_SHEET]

    summary_ws = workbook.create_sheet(SUMMARY_SHEET)

    headers = [
        "No.",
        "Student ID",
        "Name",
        "Exercise Raw",
        "Exercise %",
        "Exercise 70%",
        "Attendance",
        "Attendance %",
        "Attendance 30%",
        "Grand Total",
        "Performance",
    ]

    summary_ws.append(headers)

    for index, result in enumerate(results, start=1):
        summary_ws.append([
            index,
            result["student_id"],
            #result["student_name"],
            "----",
            (
                f'{result["exercise_raw"]:.2f}/'
                f'{result["exercise_possible"]:.2f}'
            ),
            result["exercise_percent"],
            result["exercise_weighted"],
            (
                f'{result["attended"]}/'
                f'{result["attendance_days"]}'
            ),
            result["attendance_percent"],
            result["attendance_weighted"],
            result["grand_total"],
            result["stars"],
        ])

    # --------------------------------------------------------
    # Format Summary worksheet
    # --------------------------------------------------------

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
        size=9
    )

    body_font = Font(size=8)

    thin_border = Border(
        left=Side(style="thin", color="D9E1F2"),
        right=Side(style="thin", color="D9E1F2"),
        top=Side(style="thin", color="D9E1F2"),
        bottom=Side(style="thin", color="D9E1F2"),
    )

    for cell in summary_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.border = thin_border

    for row in summary_ws.iter_rows(
        min_row=2,
        max_row=summary_ws.max_row
    ):
        for cell in row:
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=False
            )

    # Center numeric and evaluation columns
    for row in range(2, summary_ws.max_row + 1):
        for col in [1, 2, 4, 5, 6, 7, 8, 9, 10, 11]:
            summary_ws.cell(row, col).alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

    # Number formats
    for row in range(2, summary_ws.max_row + 1):
        summary_ws.cell(row, 5).number_format = "0.00"
        summary_ws.cell(row, 6).number_format = "0.00"
        summary_ws.cell(row, 8).number_format = "0.00"
        summary_ws.cell(row, 9).number_format = "0.00"
        summary_ws.cell(row, 10).number_format = "0.00"

    column_widths = {
        "A": 6,
        "B": 15,
        "C": 29,
        "D": 14,
        "E": 11,
        "F": 13,
        "G": 12,
        "H": 12,
        "I": 14,
        "J": 13,
        "K": 13,
    }

    for column, width in column_widths.items():
        summary_ws.column_dimensions[column].width = width

    summary_ws.freeze_panes = "A2"
    summary_ws.auto_filter.ref = (
        f"A1:K{summary_ws.max_row}"
    )
    summary_ws.sheet_view.showGridLines = False

    # Add summary information below the table
    information_row = summary_ws.max_row + 3

    summary_ws.cell(
        information_row,
        1,
        "Exercise maximum score"
    )
    summary_ws.cell(
        information_row,
        2,
        total_exercise_possible
    )

    summary_ws.cell(
        information_row + 1,
        1,
        "Exercise weight"
    )
    summary_ws.cell(
        information_row + 1,
        2,
        EXERCISE_WEIGHT / 100
    )
    summary_ws.cell(
        information_row + 1,
        2
    ).number_format = "0%"

    summary_ws.cell(
        information_row + 2,
        1,
        "Attendance sessions"
    )
    summary_ws.cell(
        information_row + 2,
        2,
        total_attendance_days
    )

    summary_ws.cell(
        information_row + 3,
        1,
        "Attendance weight"
    )
    summary_ws.cell(
        information_row + 3,
        2,
        ATTENDANCE_WEIGHT / 100
    )
    summary_ws.cell(
        information_row + 3,
        2
    ).number_format = "0%"

    summary_ws.cell(
        information_row + 5,
        1,
        "Total students"
    )
    summary_ws.cell(
        information_row + 5,
        2,
        total_students
    )

    summary_ws.cell(
        information_row + 6,
        1,
        "Mean Grand Total"
    )
    summary_ws.cell(
        information_row + 6,
        2,
        mean_score
    )
    summary_ws.cell(
        information_row + 6,
        2
    ).number_format = "0.00"

    summary_ws.cell(
        information_row + 7,
        1,
        "SD Grand Total (population)"
    )
    summary_ws.cell(
        information_row + 7,
        2,
        standard_deviation
    )
    summary_ws.cell(
        information_row + 7,
        2
    ).number_format = "0.00"

    for row in range(information_row, information_row + 8):
        summary_ws.cell(row, 1).font = Font(size=9, bold=True)
        summary_ws.cell(row, 2).font = Font(size=9)

    # Embed the PNG histogram on the right side of the Summary sheet.
    histogram_image = ExcelImage(OUTPUT_HISTOGRAM)
    histogram_image.width = 800
    histogram_image.height = 464
    summary_ws.add_image(histogram_image, "M2")

    workbook.save(OUTPUT_XLSX)

    # --------------------------------------------------------
    # Create compact GitHub Markdown
    # --------------------------------------------------------

    current_date = f"{date.today().day} {date.today().strftime('%b %Y')}"

    markdown_lines = [
        "# Class Performance Summary",
        "",
        (
            f"<sub>Exercise: {EXERCISE_WEIGHT:.0f}% · "
            f"Attendance: {ATTENDANCE_WEIGHT:.0f}% · "
            f"Exercise maximum: {total_exercise_possible:g} · "
            f"Attendance sessions: {total_attendance_days}</sub>"
        ),
        "",
        (
            f"**Grand Total statistics:** "
            f"Students = {total_students} · "
            f"Mean = {mean_score:.2f} · "
            f"SD = {standard_deviation:.2f} (population)"
        ),
        "",
        (
            f'## <span style="color:red">'
            f'Score Distribution until {current_date} !!!</span>'
        ),
        "",
        "![Grand Total score histogram](Grand_Total_Histogram.png)",
        "",
        (
            "| <sub>No.</sub> "
            "| <sub>Student ID</sub> "
            "| <sub>Name</sub> "
            "| <sub>Exercise</sub> "
            "| <sub>Exercise 70%</sub> "
            "| <sub>Attendance</sub> "
            "| <sub>Attendance 30%</sub> "
            "| <sub>Grand Total</sub> "
            "| <sub>Performance</sub> |"
        ),
        (
            "|---:|:---:|:---|---:|---:|:---:|---:|---:|:---:|"
        ),
    ]

    for index, result in enumerate(results, start=1):
        exercise_display = (
            f'{result["exercise_raw"]:.2f}/'
            f'{result["exercise_possible"]:.2f}'
        )

        attendance_display = (
            f'{result["attended"]}/'
            f'{result["attendance_days"]}'
        )

        markdown_lines.append(
            "| "
            + " | ".join([
                markdown_text(index),
                markdown_text(result["student_id"]),
                #markdown_text(result["student_name"]),
                '----',
                markdown_text(exercise_display),
                markdown_text(
                    f'{result["exercise_weighted"]:.2f}'
                ),
                markdown_text(attendance_display),
                markdown_text(
                    f'{result["attendance_weighted"]:.2f}'
                ),
                markdown_text(
                    f'**{result["grand_total"]:.2f}**'
                ),
                markdown_text(result["stars"]),
            ])
            + " |"
        )

    markdown_lines.extend([
        "",
        "<sub>Performance rating: "
        "★★★★★ ≥ 90, ★★★★☆ ≥ 80, ★★★☆☆ ≥ 70, "
        "★★☆☆☆ ≥ 60, ★☆☆☆☆ &lt; 60.</sub>",
        "",
    ])

    OUTPUT_MD.write_text(
        "\n".join(markdown_lines),
        encoding="utf-8"
    )

    print("Completed successfully")
    print(f"Students              : {len(results)}")
    print(f"Exercise columns      : "
          f"{', '.join(get_column_letter(c) for c in exercise_columns)}")
    print(f"Exercise maximum      : {total_exercise_possible:g}")
    print(f"Attendance columns    : "
          f"{', '.join(get_column_letter(c) for c in attendance_columns)}")
    print(f"Attendance sessions   : {total_attendance_days}")
    print(f"Grand Total mean      : {mean_score:.2f}")
    print(f"Grand Total SD        : {standard_deviation:.2f}")
    print(f"Excel output          : {OUTPUT_XLSX.resolve()}")
    print(f"Markdown output       : {OUTPUT_MD.resolve()}")
    print(f"Histogram output      : {OUTPUT_HISTOGRAM.resolve()}")


if __name__ == "__main__":
    main()
